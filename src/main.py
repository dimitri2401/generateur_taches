#################################################
# Générateur de tâches - v0.1
#################################################

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Border, Side
import numpy as np
import pulp as pl
import yaml
import os

#################################################
# Chargement des paramètres depuis le fichier YAML
#################################################

# Charger le fichier de configuration
with open('config.yaml', 'r', encoding='utf-8') as f:
    config = yaml.safe_load(f)

# Noms des fichiers Excel
input_file = config['files']['input']
output_file = config['files']['output']

# Paramètres d'optimisation

# Facteurs à inclure dans la fonction objectif
hours_enabled = config['optimization']['hours_enabled']
liberation_enabled = config['optimization']['liberation_enabled']
min_cours_enabled = config['optimization']['min_cours_enabled']
min_prof_enabled = config['optimization']['min_prof_enabled']
if min_cours_enabled or min_prof_enabled:
    nbr_min_count = config['optimization']['nbr_min_count']

if min_cours_enabled and min_prof_enabled:
    print("Les options min_cours_enabled et min_prof_enabled ne doivent pas être toutes deux activées.")
    print("Veuillez corriger le fichier de configuration et réessayer.")
    quit()

allow_negative_preferences = config['optimization']['allow_negative_preferences']
enforce_positive_with_negative = config['optimization']['enforce_positive_with_negative']
max_hours = config['optimization']['max_hours']

# Paramètres pour la génération de tâches alternatives
epsilon_strategy = config['taches_alternatives']['epsilon_strategy']
epsilon_value = config['taches_alternatives']['epsilon_value']
nbr_taches_alteratives = config['taches_alternatives']['nbr_taches_alternatives']

# Paramètres du solveur
solver_time_limit = config['solver']['time_limit']
print_solver_log = config['solver']['afficher_log']
save_solver_log = config['solver']['sauvegarder_log']
if save_solver_log:
    solver_log_file = config['solver']['fichier_log']
else:
    solver_log_file = None

#################################################
# Chargement des données depuis le fichier Excel
#################################################

# Nom du fichier Excel contenant les données de la tâche
tache_file = input_file

def load_data(tache_file):
    # Charge les feuilles du fichier Excel
    dfs = pd.read_excel(tache_file, sheet_name=["PROF","COURS","PREF","MAX_NB_GR","ATTRIB_PREALABLE"], header=None)

    df_prof = dfs["PROF"]
    df_cours = dfs["COURS"]
    df_pref = dfs["PREF"]
    df_pref.fillna(0, inplace=True)
    df_max_nb_gr = dfs["MAX_NB_GR"]
    df_max_nb_gr.fillna(0, inplace=True)
    df_attrib_preal = dfs["ATTRIB_PREALABLE"]
    df_attrib_preal.fillna(0, inplace=True)

    # Extraction des données de la page PROF
    nbr_prof = int(df_prof.iat[0, 0])
    
    # Extraire toutes les colonnes en une fois
    list_prof = df_prof.iloc[1:nbr_prof+1, 0].to_numpy(dtype=str)
    liberation_prof = df_prof.iloc[1:nbr_prof+1, 1].to_numpy(dtype=float)
    ci_cible = df_prof.iloc[1:nbr_prof+1, 2].to_numpy(dtype=float)
    
    # Calcul de ci_min et ci_max
    ci_min = (df_prof.iloc[1:nbr_prof+1, 2] - df_prof.iloc[1:nbr_prof+1, 3]).to_numpy(dtype=float)
    ci_max = (df_prof.iloc[1:nbr_prof+1, 2] + df_prof.iloc[1:nbr_prof+1, 4]).to_numpy(dtype=float)
    
    prep_min = df_prof.iloc[1:nbr_prof+1, 5].to_numpy(dtype=int)  # Non utilisé pour l'instant
    prep_max = df_prof.iloc[1:nbr_prof+1, 6].to_numpy(dtype=int)
    gr_max = df_prof.iloc[1:nbr_prof+1, 7].to_numpy(dtype=int)
    ci_ant = df_prof.iloc[1:nbr_prof+1, 10].to_numpy(dtype=float)

    # Extraction des données de la page COURS
    nbr_cours = int(df_cours.iat[0, 0])
    
    list_cours = df_cours.iloc[1:nbr_cours+1, 0].to_numpy(dtype=str)
    periodes_cours = df_cours.iloc[1:nbr_cours+1, 1].to_numpy(dtype=int)
    groupes_cours = df_cours.iloc[1:nbr_cours+1, 2].to_numpy(dtype=int)
    etudiants_cours = df_cours.iloc[1:nbr_cours+1, 3].to_numpy(dtype=int)

    # Extraction des préférences, nombres max de groupes et attributions préalables
    pref_prof = df_pref.iloc[1:nbr_prof+1, 1:nbr_cours+1].to_numpy(dtype=int) # Mettre float si on veut accepter les préférences réelles
    max_nb_gr = df_max_nb_gr.iloc[1:nbr_prof+1, 1:nbr_cours+1].to_numpy(dtype=int)
    attrib_preal = df_attrib_preal.iloc[1:nbr_prof+1, 1:nbr_cours+1].to_numpy(dtype=int)

    return {
        'professors': {
            'list': list_prof,
            'nbr': nbr_prof,
            'liberation': liberation_prof,
            'ci_cible': ci_cible,
            'ci_min': ci_min,
            'ci_max': ci_max,
            'prep_min': prep_min,
            'prep_max': prep_max,
            'gr_max': gr_max,
            'ci_ant': ci_ant
        },
        'courses': {
            'list': list_cours,
            'nbr': nbr_cours,
            'periodes': periodes_cours,
            'groupes': groupes_cours,
            'etudiants': etudiants_cours
        },
        'preferences': {
            'prof': pref_prof,
            'max_nb_gr': max_nb_gr,
            'attrib_preal': attrib_preal
        }
    }

data = load_data(tache_file)

nbr_prof = data['professors']['nbr']
list_prof = data['professors']['list']
liberation_prof = data['professors']['liberation']
ci_cible = data['professors']['ci_cible']
ci_min = data['professors']['ci_min']
ci_max = data['professors']['ci_max']
prep_min = data['professors']['prep_min']
prep_max = data['professors']['prep_max']
gr_max = data['professors']['gr_max']
ci_ant = data['professors']['ci_ant']

nbr_cours = data['courses']['nbr']
list_cours = data['courses']['list']
periodes_cours = data['courses']['periodes']
groupes_cours = data['courses']['groupes']
etudiants_cours = data['courses']['etudiants']

pref_prof = data['preferences']['prof']
max_nb_gr = data['preferences']['max_nb_gr']
attrib_preal = data['preferences']['attrib_preal']

# Liste les indices des attributions préalables
attrib_preal_indices = [(i, j, attrib_preal[i][j]) for i in range(nbr_prof) for j in range(nbr_cours) if attrib_preal[i][j] != 0]

#################################################
# Définition du problème de programmation linéaire
#################################################

# Pré-calcul des préférences pénalisées au carré pour tous les (i,j) pour éviter les recalculs inutiles
pref_matrix = np.array(1 - (2 - pref_prof)**2/12 - (2 - pref_prof)/6, dtype=float)

# Retourne la préférence pénalisée au carré pré-calculée du prof i pour le cours j
def pref(i, j):
    return pref_matrix[i, j]

# Création du problème
prob = pl.LpProblem("probleme_tache", pl.LpMaximize)

# Variables de décision y[i][j]: le cours j est-il attribué au prof i (1 si oui, 0 sinon)
y = pl.LpVariable.dicts("y", ((i,j) for i in range(nbr_prof) for j in range(nbr_cours)), cat='Binary')

# Variables de décision x[i][j]: nombre de groupes du cours j attribués au prof i
x = pl.LpVariable.dicts("x", ((i,j) for i in range(nbr_prof) for j in range(nbr_cours)), lowBound=0, cat='Integer')

# Construction de la fonction objectif
def base_obj_term(i, j):
    term = x[(i,j)] * pref(i,j) # Fonction de base à optimiser : nbr de groupes * préférence
    if hours_enabled:
        term = periodes_cours[j] * term # Pondération par le nombre de périodes du cours
    if liberation_enabled:
        if liberation_prof[i] != 0: # Éviter la division par zéro
            term = term / liberation_prof[i] # Prendre en compte la libération du prof
    return term

# Compter plusieurs fois la préférence minimum si activé
if min_cours_enabled or min_prof_enabled:
    # min_pref_value représente la préférence minimale parmi les cours ou les profs
    min_pref_value = pl.LpVariable("min_pref_value", lowBound=None) # Borne dépendant des options activées - Difficile de mesurer sa lowBound
    
    # min_pref_value est la pire des préférences parmi l'ensemble des cours attribués
    if min_cours_enabled:
        for i in range(nbr_prof):
            for j in range(nbr_cours):
                prob += min_pref_value <= base_obj_term(i, j)

    # min_pref_value est la pire des préférences parmi l'ensemble des professeurs
    if min_prof_enabled:
        for i in range(nbr_prof):
            prob += min_pref_value <= pl.lpSum([base_obj_term(i, j) for j in range(nbr_cours)])

    # La fonction objectif inclut le minimum un certain nombre de fois puis la somme totale des termes
    objective_fct = nbr_min_count * min_pref_value + pl.lpSum([base_obj_term(i, j) for i in range(nbr_prof) for j in range(nbr_cours)])
else:
    # La fonction objectif est la somme des termes
    objective_fct = pl.lpSum([base_obj_term(i, j) for i in range(nbr_prof) for j in range(nbr_cours)])

# Ajouter la fonction objectif au problème
prob += objective_fct

# Contrainte d'optimisation: nombre maximum de groupes par prof pour chaque cours
for i in range(nbr_prof):
    for j in range(nbr_cours):
        prob += x[(i,j)] >= y[(i,j)]
        prob += x[(i,j)] <= y[(i,j)] * max_nb_gr[i][j]

# Contrainte d'optimisation: tous les groupes de chaque cours doivent être attribués
for j in range(nbr_cours):
    prob += pl.lpSum([x[(i,j)] for i in range(nbr_prof)]) == groupes_cours[j]

# hc[i] = heures de cours pour le prof i
hc = {i: pl.lpSum([periodes_cours[j] * x[(i,j)] for j in range(nbr_cours)]) for i in range(nbr_prof)}

# hp[i] = heures de préparation pour le prof i
hp = {i: pl.lpSum([periodes_cours[j] * y[(i,j)] for j in range(nbr_cours)]) for i in range(nbr_prof)}

# nes[i] = nombre d'étudiants pour le prof i
nes = {i: pl.lpSum([etudiants_cours[j] * x[(i,j)] for j in range(nbr_cours)]) for i in range(nbr_prof)}

# pes[i] = périodes × étudiants pour le prof i
pes = {i: pl.lpSum([periodes_cours[j] * x[(i,j)] * etudiants_cours[j] for j in range(nbr_cours)]) for i in range(nbr_prof)}

# nbr_prep[i] = nombre de préparations pour le prof i
nbr_prep = {i: pl.lpSum([y[(i,j)] for j in range(nbr_cours)]) for i in range(nbr_prof)}

# Calcul de la CI

# Variables binaires z[i]: z[i]=1 si nes(i) >= 75, 0 sinon (essentielles pour la linéarisation des conditionnelles dans le calcul de la CI)
z = pl.LpVariable.dicts("z", (i for i in range(nbr_prof)), cat='Binary')
# Variables entières w[i]: w[i]=nes(i) si nes(i) >= 75, 0 sinon (essentielles pour la linéarisation des conditionnelles dans le calcul de la CI)
w = pl.LpVariable.dicts("w", (i for i in range(nbr_prof)), lowBound=0, cat='Integer')

# Variables binaires t[i]: t[i]=1 si pes(i) >= 416, 0 sinon (essentielles pour la linéarisation des conditionnelles dans le calcul de la CI)
t = pl.LpVariable.dicts("t", (i for i in range(nbr_prof)), cat='Binary')
# Variables entières m[i]: m[i]=pes(i)-415 si pes(i) >= 416, 0 sinon (essentielles pour la linéarisation des conditionnelles dans le calcul de la CI)
m = pl.LpVariable.dicts("m", (i for i in range(nbr_prof)), lowBound=0, cat='Integer')

# Big M (pour la linéarisation des contraintes conditionnelles dans la CI) - Valeurs à ajuster si nécessaire

M1 = 200 # >= à la valeur maximale possible de nes(i)
# Contraintes pour w[i] = nes[i] si nes[i] >= 75, sinon w[i] = 0 avec la méthode Big M
for i in range(nbr_prof):
    # w[i] = nes[i] si z[i]=1, sinon w[i]=0
    prob += w[(i)] <= M1 * z[(i)]  # Force w[i]=0 si z[i]=0
    prob += w[(i)] >= nes[i] - M1 * (1 - z[(i)])  # Force w[i]=nes[i] si z[i]=1
    prob += w[(i)] <= nes[i] + M1 * (1 - z[(i)])  # Force w[i]=nes[i] si z[i]=1
    # z[i]=1 si nes[i] >= 75
    prob += nes[i] <= 74 + M1 * z[(i)]  # Si nes[i] >= 75, alors z[i]=1
    #prob += nes[i] >= 75 - M1 * (1 - z[(i)])  # Si nes[i] < 75, alors z[i]=0 - Inutile, l'optimiseur ne cherchera pas à augmenter nes[i] inutilement

M2 = 400 # >= à la valeur maximale possible de pes[i] - 415 - Peut probablement être réduit
# Contraintes pour m[i] = max(0, pes[i]-415) par la méthode Big M
for i in range(nbr_prof):
    # m[i] représente l'excédant de pes[i] au-delà de 415
    prob += m[(i)] >= pes[i] - 415
    #prob += m[(i)] >= 0 # Inutile, car m[i] est déjà défini avec lowBound=0
    prob += m[(i)] <= M2 * t[(i)]
    prob += m[(i)] <= pes[i] - 415 + M2 * (1 - t[(i)])

# Pré-calcul de la CI pour chaque prof
ci = {i: hc[i] * 1.2 + hp[i] * 0.9 + pes[i] * 0.04 + m[(i)] * 0.03 + w[(i)] * 0.01 
      for i in range(nbr_prof)}
# On suppose que nes ne dépasse pas 160 pour éviter la pénalité au carré problématiques pour la linéarisation
# Pour qu'il dépasse 160, il faudrait au moins 4 groupes de 40 étudiants, ce qui ne serait de toute façon pas accepté dans nos scénarios

# CI des profs entre min et max
for i in range(nbr_prof):
    prob += ci[i] >= ci_min[i]
    prob += ci[i] <= ci_max[i]

# Nbr d'heures entre 0 et max_hours (défini dans config.yaml)
for i in range(nbr_prof):
    prob += hc[i] <= max_hours

# Nbr total de groupe max par prof entre 0 et gr_max
for i in range(nbr_prof):
    prob += pl.lpSum([x[(i,j)] for j in range(nbr_cours)]) <= gr_max[i]

# Nbr de préparations max par prof
for i in range(nbr_prof):
    prob += nbr_prep[i] <= prep_max[i]

# Attribution préalables
for i, j, nb_groupes in attrib_preal_indices:
    prob += y[(i,j)] == 1
    prob += x[(i,j)] >= nb_groupes

# Contrainte: aucun cours négatif
if not allow_negative_preferences:
    for i in range(nbr_prof):
        for j in range(nbr_cours):
            prob += pref_prof[i,j] * y[(i,j)] >= 0

# Contrainte: si un professeur reçoit une préférence négative pour un cours, s'assurer qu'il ait un positif au moins aussi grand à côté
if allow_negative_preferences and enforce_positive_with_negative:
    for i in range(nbr_prof):
        prob += pl.lpSum([y[(i,j)] * pref_prof[i][j] for j in range(nbr_cours)]) >= 0

#################################################
# Traitement des résultats (arrondir les valeurs qui devraient être entières pour éviter les potentielles approximations de la résolution selon le solver utilisé)
# Probablement un peu inutile mais assure un résultat valide
#################################################

def get_val(v):# Retourne v arrondi à l'entier, ou 0 si None - Au cas où la valeur retournée par le solver serait None
    return 0 if v.varValue is None else round(v.varValue)

def round_xy(): # Stocke dans un np.array les valeurs arrondies des variables de décision
    global y_output, x_output
    y_output = np.array([[get_val(y[(i,j)]) for j in range(nbr_cours)] for i in range(nbr_prof)])
    x_output = np.array([[get_val(x[(i,j)]) for j in range(nbr_cours)] for i in range(nbr_prof)])

def hc_output(i):
    # sum_j of periodes_cours[j] * x_output[i,j]
    return int(np.dot(periodes_cours, x_output[i]))

def hp_output(i):
    # sum_j periodes_cours[j] * y_output[i,j]
    return int(np.dot(periodes_cours, y_output[i]))

def nes_output(i):
    # sum_j etudiants_cours[j] * x_output[i,j]
    return int(np.dot(etudiants_cours, x_output[i]))

coeff = periodes_cours * etudiants_cours # Pré-calcul des coefficients pour pes_output
def pes_output(i):
    # sum_j (periodes[j] * etudiants[j] * x_output[i,j])
    return int(np.dot(coeff, x_output[i]))

def nbr_prep_output(i):
    # sum_j y_output[i,j]
    return int(np.sum(y_output[i]))

def ci_output(i):
    if pes_output(i) >= 416:
        total = hc_output(i) * 1.2 + hp_output(i) * 0.9 + 415 * 0.04 + (pes_output(i)-415) * 0.07
    else:
        total = hc_output(i) * 1.2 + hp_output(i) * 0.9 + pes_output(i) * 0.04
    if nes_output(i) >= 75:
        total += nes_output(i) * 0.01
    # On suppose que nes ne dépasse pas 160 pour éviter la pénalité au carré qu'il faudrait linéariser
    return total

#################################################
# Exportation des résultats dans un fichier Excel
#################################################

def save_excel_file(writer, numero_tache): # Exportation des résultats dans un fichier Excel comme nouvelle feuille
    round_xy()  # Arrondir les valeurs des variables de décision

    # Création du DataFrame des résultats
    data = {}
    df_tache = pd.DataFrame(data)

    # Remplissage du DataFrame avec les résultats
    # Nom des profs
    df_tache['Prof'] = list_prof

    # Nombre de groupes attribués par cours
    for j in range(nbr_cours):
        col_name = str(list_cours[j])
        df_tache[col_name] = [x_output[i,j] for i in range(nbr_prof)]

    # Libération
    df_tache['% cours'] = liberation_prof 

    # CI des profs après attribution
    ci_values = []
    ci_pleine_values = []
    ci_annee_values = []

    # Pénalité de préférence
    pref_pond_heure_values = []

    # Autres indicateurs
    nbr_prep_values = []
    nbr_gr_values = []
    nbr_heures_values = []

    for i in range(nbr_prof):
        ci_values.append(ci_output(i))
        ci_pleine_values.append(ci_output(i) + 40 * (1-liberation_prof[i]))
        ci_annee_values.append(ci_ant[i] + ci_pleine_values[i])
        nbr_prep_values.append(nbr_prep_output(i))
        nbr_gr_values.append(sum([x_output[i,j] for j in range(nbr_cours)]))
        nbr_heures_values.append(hc_output(i))
        if nbr_heures_values[i] == 0:
            pref_pond_heure_values.append(0)
        else:
            pref_pond_heure_values.append(round(sum(periodes_cours[j] * x_output[i,j] * pref(i,j)/nbr_heures_values[i] for j in range(nbr_cours)),2))

    df_tache['CI Cours'] = ci_values
    df_tache['CI Pleine'] = ci_pleine_values
    df_tache['CI Année'] = ci_annee_values
    df_tache['Nbr. Prép.'] = nbr_prep_values
    df_tache['Nbr. Groupes'] = nbr_gr_values
    df_tache['Nbr. Périodes'] = nbr_heures_values
    df_tache['Préf.'] = pref_pond_heure_values

    # Nom de la feuille. "1" pour l'optimale, "2", "3", etc. pour les alternatives
    sheet_name = str(numero_tache + 1)

    # Mise en forme et enregistrement du fichier Excel
    df_tache.to_excel(writer, sheet_name=sheet_name, index=False)
    sheet = writer.sheets[sheet_name]

    # Coloration des cellules contenant des préférences négatives
    for i in range(nbr_prof):
        for j in range(nbr_cours):
            if pref_prof[i][j] == -1 and y_output[i,j] == 1:
                cell = sheet.cell(row=i+2, column=j+2)  # +2 à cause des en-têtes
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            if pref_prof[i][j] == -2 and y_output[i,j] == 1:
                cell = sheet.cell(row=i+2, column=j+2)  # +2 à cause des en-têtes
                cell.fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")

    # Styles de bordure pour les cellules précédentes (les cellules des statiques seront traitées séparément)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin'))
        
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

    # Colonne (% cours) en % avec 1 décimale
    col_liberation = df_tache.columns.get_loc('% cours') + 1
    for row in range(2, nbr_prof + 2):
        cell = sheet.cell(row=row, column=col_liberation)
        cell.number_format = '0.0%'

    # Ajout du résumé sous le tableau
    resume_row = nbr_prof + 2  # Première ligne après les données des profs
    
    # Titres des cours
    for j in range(nbr_cours):
        cell = sheet.cell(row=resume_row, column=j+2, value=str(list_cours[j]))
        cell.border = thin_border

    # "Commande" - nombre de groupes à assigner par cours
    sheet.cell(row=resume_row+1, column=1, value="Commande").border = thin_border
    for j in range(nbr_cours):
        cell = sheet.cell(row=resume_row+1, column=j+2, value=int(groupes_cours[j]))
        cell.border = thin_border
    
    # "Assignés" - nombre de groupes réellement assignés par cours
    sheet.cell(row=resume_row+2, column=1, value="Assignés").border = thin_border
    for j in range(nbr_cours):
        assigned = int(sum(x_output[i,j] for i in range(nbr_prof)))
        cell = sheet.cell(row=resume_row+2, column=j+2, value=assigned)
        cell.border = thin_border
    
    # Ligne "Nbr. Étud." - nombre d'étudiants par cours
    sheet.cell(row=resume_row+3, column=1, value="Nbr. Étud.").border = thin_border
    for j in range(nbr_cours):
        cell = sheet.cell(row=resume_row+3, column=j+2, value=etudiants_cours[j])
        cell.border = thin_border
    
    # Ligne "Per/sem" - nombre de périodes par semaine par cours
    sheet.cell(row=resume_row+4, column=1, value="Pér./sem.").border = thin_border
    for j in range(nbr_cours):
        cell = sheet.cell(row=resume_row+4, column=j+2, value=int(periodes_cours[j]))
        cell.border = thin_border

    # Ajout des statistiques sur les préférences (2 colonnes à droite du tableau principal)
    stats_col = len(df_tache.columns) + 2
    
    # Statistiques sous forme de liste
    stats = [
        ("Moy. pénal.", round(np.mean(pref_pond_heure_values), 2)),
        ("É.-t. pénal.", round(np.std(pref_pond_heure_values), 2)),
        ("Min. préf", round(np.min(pref_pond_heure_values), 2)),
        ("", ""),  # Ligne vide
        ("Prof à 1 prép", sum(1 for val in nbr_prep_values if val == 1)),
        ("Prof à 4 gr.", sum(1 for val in nbr_gr_values if val >= 4)),
        ("Prof à 15 h.", sum(1 for val in nbr_heures_values if val >= 15)),
        ("Tot. satisfaits", sum(1 for val in pref_pond_heure_values if val == 1)),
        ("", ""),  # Ligne vide
        ("CI Moy.", round(np.mean(ci_pleine_values), 2)),
        ("CI É.-t.", round(np.std(ci_pleine_values), 2)),
        ("CI Min", round(np.min(ci_pleine_values), 2)),
        ("CI Max", round(np.max(ci_pleine_values), 2)),
        ("sCI/40", round(sum(ci_values)/40, 2)),
        ("CI Année Moy.", round(np.mean(ci_annee_values), 2)),
        ("CI Année É.-t.", round(np.std(ci_annee_values), 2)),
        ("CI Année Min", round(np.min(ci_annee_values), 2)),
        ("CI Année Max", round(np.max(ci_annee_values), 2)),
    ]
    
    # Écrire les statistiques dans la feuille Excel
    for i, (label, value) in enumerate(stats, start=2):
        if label:  # Ignorer les lignes vides
            sheet.cell(row=i, column=stats_col, value=label).border = thin_border
            sheet.cell(row=i, column=stats_col + 1, value=value).border = thin_border

    # Centrer le contenu de la feuille de calcul
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Écrire les titres à la verticale (en-tête)
    for cell in sheet[1]:
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 80
    
    # Écrire les titres à la verticale (résumé)
    for j in range(nbr_cours):
        cell = sheet.cell(row=resume_row, column=j+2)
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
    sheet.row_dimensions[resume_row].height = 60

    # Auto-ajustement de la largeur des colonnes (exclure la première ligne avec texte vertical)
    for column in sheet.columns:
        column_letter = column[0].column_letter
        # Calculer la largeur basée sur les données (lignes 2+), pas les titres verticaux
        max_length = max(
            (len(str(cell.value)) for cell in column[1:] if cell.value),
            default=0
        )
        # Limiter la largeur: minimum 3, maximum 12 pour éviter les colonnes trop larges
        sheet.column_dimensions[column_letter].width = max(min(max_length + 2, 15), 8)
    
    # Affiche la préférence moyenne et min dans le terminal
    print()
    print(f"Tâche {numero_tache + 1}")
    print(f"Préférence moyenne: {np.mean(pref_pond_heure_values):.4f}")
    print(f"Préférence minimum: {np.min(pref_pond_heure_values):.2f}")
    print()


#################################################
# Résolution du problème de programmation linéaire
#################################################

# Sélection du solveur - Highs utilisé par défaut - Attention à installer les dépendances nécessaires, python -m pip install pulp[highs]
solver = pl.HiGHS(timeLimit=solver_time_limit, log_to_console = print_solver_log, log_file = solver_log_file)

# Vérifier si le fichier de sortie existe déjà et numéroter le nouveau si nécessaire
result_file = os.path.join(os.getcwd(), output_file + '.xlsx')
nbr_file = 1
while os.path.exists(result_file):
    result_file = os.path.join(os.getcwd(), f"{output_file}_{nbr_file}{'.xlsx'}")
    nbr_file += 1

with pd.ExcelWriter(result_file, engine="openpyxl") as writer:
    # Résolution de la tâche optimale
    prob.solve(solver)
    save_excel_file(writer, 0)  # Feuille "1"
    
    # Génération des tâches alternatives si demandé
    if epsilon_strategy:
        for k in range(nbr_taches_alteratives):
            # Contrainte epsilon pour générer une nouvelle solution différente (mais moins bonne)
            eps_constr = (objective_fct <= pl.value(prob.objective) - (k + 1) * epsilon_value)
            prob += eps_constr
            
            # Résolution de la tâche alternative
            prob.solve(solver)

            # Suppression de la contrainte epsilon pour la prochaine itération
            prob.constraints.pop(eps_constr.name, None)
            
            # Exportation de la solution alternative comme nouvelle feuille
            save_excel_file(writer, k + 1)  # Feuille "2", "3", etc.